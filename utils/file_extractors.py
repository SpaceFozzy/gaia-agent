import os
import logging
import torch
from huggingface_hub import snapshot_download
from docx import Document
from docx.text.paragraph import Paragraph
from docx.table import Table
from openpyxl import load_workbook
from openpyxl.styles.colors import RGB
from transformers import pipeline

logger = logging.getLogger(__name__)

supported_file_types = ["docx", "xlsx", "py", "mp3"]

device = "cuda:0" if torch.cuda.is_available() else "cpu"
whisper = pipeline(
    "automatic-speech-recognition",
    "openai/whisper-small",
    chunk_length_s=30,
    device=device
)


class FileExtractor:
    def __init__(self, file_name):
        directory = os.path.join(os.path.dirname(__file__), "../", "downloaded_files")
        os.makedirs(directory, exist_ok=True)
        snapshot_download(
            repo_id="gaia-benchmark/GAIA", repo_type="dataset", local_dir=directory
        )

        self.file_name = file_name
        self.file_path = os.path.join(directory, "2023/validation", self.file_name)

        logger.info(f"Creating file extractor for {file_name}")
        is_supported = self.is_file_supported()
        if not is_supported:
            raise Exception(
                f"Unable to parse file {file_name}: file type not supported"
            )

    def get_extension(self):
        return os.path.splitext(self.file_name)[1].lstrip(".").lower()

    def is_file_supported(self):
        extension = self.get_extension()
        logger.info(f"Checking for file support for {extension}...")

        if extension in supported_file_types:
            logger.info("File type is supported.")
            return True

        logger.info("File type is not supported.")
        return False

    def iter_block_items(self, parent):
        for child in parent.element.body:
            if child.tag.endswith("p"):
                yield Paragraph(child, parent)
            elif child.tag.endswith("tbl"):
                yield Table(child, parent)

    def docx_to_text(self):
        doc = Document(self.file_path)
        lines = []
        for block in self.iter_block_items(doc):
            if isinstance(block, Paragraph):
                style = block.style.name
                text = "".join(run.text for run in block.runs).strip()
                if not text:
                    continue
                if style.startswith("Heading"):
                    lines.append(f"# {text}")
                else:
                    lines.append(text)
            elif isinstance(block, Table):
                for row in block.rows:
                    cells = [cell.text.strip().replace("\n", " ") for cell in row.cells]
                    lines.append("| " + " | ".join(cells) + " |")
        return "\n".join(lines)

    def xlsx_to_text(self):
        wb = load_workbook(self.file_path, data_only=True)
        lines = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            lines.append(f"# Sheet: {sheet}")
            for row in ws.iter_rows(values_only=False):
                if all(cell is None for cell in row):
                    continue

                row_text = ""
                for cell in row:
                    if cell is not None:
                        row_text += str(cell.value)
                        # If the cell had a color other than the default background, include
                        # it with the cell value to the LLM can be aware.
                        # Sometimes the colors are an RGB class which we will ignore for now.
                        if cell.fill.start_color.rgb != "00000000" and not isinstance(
                            cell.fill.start_color.rgb, RGB
                        ):
                            row_text += f" (#{cell.fill.start_color.rgb})"
                        row_text += " | "
                    else:
                        row_text += " | "

                lines.append(row_text)
        return "\n".join(lines)

    def mp3_to_text(self):
        result = whisper(self.file_path, return_timestamps=True)
        return result["text"]

    def raw_file_to_text(self):
        with open(self.file_path, mode="r", encoding="utf-8") as file:
            contents = file.read()
            return contents

    def extract_text(self):
        extension = self.get_extension()
        match extension:
            case "docx":
                return self.docx_to_text()
            case "xlsx":
                return self.xlsx_to_text()
            case "py":
                return self.raw_file_to_text()
            case "mp3":
                return self.mp3_to_text()
            case _:
                raise Exception(
                    f"Attempted to extract from unsupported extension {extension}"
                )

    def __call__(self):
        logger.info(f"Extracting file contents from {self.file_name}...")
        if os.path.exists(self.file_path):
            logger.info("File exists on disk.")
            text = self.extract_text()
            logger.info("Text extracted from file.")
            return text
        else:
            raise Exception("File does not exist on disk.")
